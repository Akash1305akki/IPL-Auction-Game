import json
from pathlib import Path
from typing import Optional
import pandas as pd

from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from openpyxl import load_workbook

# -------------------------------------------------------------------
# Paths & App Setup
# -------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
TEMPLATES_DIR = BASE_DIR / "templates"

EXCEL_PATH = DATA_DIR / "Auction_Sheet.xlsx"
TEAMS_JSON_PATH = DATA_DIR / "teams.json"

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))


# -----------------------------------------------------------
# Model for add-player
# -----------------------------------------------------------
class PlayerPayload(BaseModel):
    action: str
    player_name: str
    team: Optional[str] = None
    sold_for: Optional[float] = None
    speciality: Optional[str] = None
    nationality: Optional[str] = None
    base_price: Optional[float] = None
    
# -----------------------------------------------------------
# Model for fetch-tab
# -----------------------------------------------------------
class FetchTabPayload(BaseModel):
    sheet_name: str

# -----------------------------------------------------------
# Model for remove-player
# -----------------------------------------------------------
class RemovePlayerPayload(BaseModel):
    sheet_name: str
    player_name: str
    base_price: float

# -------------------------------------------------------------------
# Excel Sorting
# -------------------------------------------------------------------
def sort_all_sheets(wb):
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        if sheet_name == "Unsold":
            continue

        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            continue

        header = rows[0]
        data_rows = rows[1:]

        try:
            nat_index = header.index("Nationality")
        except ValueError:
            continue

        sorted_rows = sorted(
            data_rows,
            key=lambda r: (r[nat_index] != "Indian", r[nat_index] or "")
        )

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        for i, row_data in enumerate(sorted_rows, start=2):
            for j, value in enumerate(row_data, start=1):
                sheet.cell(row=i, column=j, value=value)


# -------------------------------------------------------------------
# Page Endpoint
# -------------------------------------------------------------------
@app.get("/auction-summary-page", response_class=HTMLResponse)
def auction_summary_page(request: Request):
    if not TEAMS_JSON_PATH.exists():
        raise HTTPException(status_code=500, detail="teams.json not found in data folder")

    with TEAMS_JSON_PATH.open("r", encoding="utf-8") as f:
        teams_data = json.load(f)

    return templates.TemplateResponse(
        "auction_summary.html",
        {
            "request": request,
            "teams": teams_data
        }
    )


# -----------------------------------------------------------
# Helper: Remove a player from UNSOLD sheet if exists
# -----------------------------------------------------------
def remove_player_from_unsold(wb, player_name: str):
    if "Unsold" not in wb.sheetnames:
        return  # Nothing to do

    sheet = wb["Unsold"]
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return

    header = rows[0]
    data_rows = rows[1:]

    # Filter out matching player
    new_rows = [row for row in data_rows if str(row[0]).strip().lower() != player_name.strip().lower()]

    # If no change → return
    if len(new_rows) == len(data_rows):
        return  

    # Clear sheet
    sheet.delete_rows(1, sheet.max_row)

    # Write header
    for col_index, value in enumerate(header, start=1):
        sheet.cell(row=1, column=col_index, value=value)

    # Write filtered data
    for r_index, row in enumerate(new_rows, start=2):
        for c_index, value in enumerate(row, start=1):
            sheet.cell(row=r_index, column=c_index, value=value)



# -------------------------------------------------------------------
# API: Add Player (Sold / Unsold)
# -------------------------------------------------------------------
@app.post("/add-player")
def add_player(payload: PlayerPayload):
    # ---------------------------
    # LOAD EXCEL
    # ---------------------------
    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=500, detail="Auction_Sheet.xlsx not found in data folder")

    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to load Excel: {e}")

    action = payload.action.strip().lower()

    # ---------------------------
    # LOAD teams.json
    # ---------------------------
    if not TEAMS_JSON_PATH.exists():
        raise HTTPException(status_code=500, detail="teams.json missing")

    with open(TEAMS_JSON_PATH, "r") as f:
        teams_data = json.load(f)

    team_key = None
    if payload.team:
        for k in teams_data.keys():
            if k.endswith(f"({payload.team})"):
                team_key = k
                break

    # ---------------------------
    # CASE 1: UNSOLD
    # ---------------------------
    if action == "unsold":
        base_price = payload.base_price if payload.base_price is not None else payload.sold_for

        if not payload.player_name:
            raise HTTPException(status_code=400, detail="Player name required for Unsold.")

        if "Unsold" not in wb.sheetnames:
            raise HTTPException(status_code=500, detail="Unsold sheet missing.")

        sheet = wb["Unsold"]
        sheet.append([payload.player_name, base_price])

    # ---------------------------
    # CASE 2: SOLD
    # ---------------------------
    elif action == "sold":

        required_fields = [
            payload.player_name,
            payload.team,
            payload.sold_for,
            payload.speciality,
            payload.nationality
        ]

        if any(v is None or v == "" for v in required_fields):
            raise HTTPException(status_code=400, detail="Sold requires all fields.")

        if payload.team not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Team sheet '{payload.team}' missing.")

        sheet = wb[payload.team]

        # Add player to team sheet
        sheet.append([
            payload.player_name,
            payload.sold_for,
            payload.speciality,
            payload.nationality
        ])

        # ---------------------------
        # REMOVE FROM UNSOLD SHEET
        # ---------------------------
        remove_player_from_unsold(wb, payload.player_name)

        # ---------------------------
        # UPDATE JSON
        # ---------------------------
        if not team_key:
            raise HTTPException(status_code=400, detail="Team not found in JSON.")

        team_json = teams_data[team_key]

        team_json["Purse Remaining"] -= payload.sold_for
        team_json["Slots Filled"] += 1

        if payload.nationality.lower() == "indian":
            team_json["Indian Slots Remaining"] -= 1
        else:
            team_json["Overseas Slots Remaining"] -= 1

        with open(TEAMS_JSON_PATH, "w") as f:
            json.dump(teams_data, f, indent=4)

    else:
        raise HTTPException(status_code=400, detail="Invalid action. Must be 'Sold' or 'Unsold'.")

    # ---------------------------
    # SORT & SAVE EXCEL
    # ---------------------------
    sort_all_sheets(wb)

    try:
        wb.save(EXCEL_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to save Excel: {e}")

    return {"status": "success", "message": "Player added successfully."}


# -----------------------------------------------------------
# Endpoint: Fetch sheet as JSON
# -----------------------------------------------------------
@app.post("/fetch-tab")
def fetch_tab(payload: dict):
    sheet_name = payload["sheet_name"]

    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=500, detail="Excel file not found")

    df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)

    # 1. Drop columns where all values are NaN
    df = df.dropna(axis=1, how="all")

    # 2. Drop rows where all values are NaN
    df = df.dropna(axis=0, how="all")

    # 3. Strip header names (Excel often adds extra spaces)
    df.columns = [str(c).strip() for c in df.columns]

    # 4. Replace NaN values with empty string
    df = df.fillna("")

    # Convert to list of JSON rows
    data = df.to_dict(orient="records")

    return data

# -----------------------------------------------------------
# Endpoint: Remove Players and Move To Unsold
# ----------------------------------------------------------
@app.post("/remove-player")
def remove_player(payload: RemovePlayerPayload):

    if not EXCEL_PATH.exists():
        raise HTTPException(status_code=500, detail="Excel file not found")

    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    sheet_name = payload.sheet_name
    player_name = payload.player_name.strip()
    base_price = payload.base_price

    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found")

    team_sheet = wb[sheet_name]

    if "Unsold" not in wb.sheetnames:
        raise HTTPException(status_code=500, detail="Unsold sheet missing")
    unsold_sheet = wb["Unsold"]

    # ---------------------------------
    # REMOVE PLAYER FROM TEAM SHEET
    # ---------------------------------
    rows = list(team_sheet.iter_rows(values_only=True))
    header = rows[0]
    new_rows = [header]
    removed = False

    for row in rows[1:]:
        if row[0] == player_name:
            removed = True
            continue
        new_rows.append(row)

    if not removed:
        raise HTTPException(
            status_code=404,
            detail=f"Player '{player_name}' not found in {sheet_name}"
        )

    # Clear and rewrite team sheet
    for r in team_sheet.iter_rows():
        for c in r:
            c.value = None

    for r_idx, row_data in enumerate(new_rows, start=1):
        for c_idx, val in enumerate(row_data, start=1):
            team_sheet.cell(row=r_idx, column=c_idx, value=val)

    # ---------------------------------
    # ADD TO UNSOLD SHEET
    # ---------------------------------
    unsold_sheet.append([player_name, base_price])

    # ---------------------------------
    # SORT ALL SHEETS (except Unsold)
    # ---------------------------------
    sort_all_sheets(wb)
    
    # SAVE EXCEL
    try:
        wb.save(EXCEL_PATH)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel save failed: {e}")

    return {
        "status": "success",
        "message": f"Player '{player_name}' moved from {sheet_name} to Unsold and sheets sorted."
    }

# -------------------------------------------------------------------
# Root
# -------------------------------------------------------------------
@app.get("/")
def root():
    return {"message": "Auction API is running!"}
